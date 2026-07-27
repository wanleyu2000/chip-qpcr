#!/usr/bin/env python3
"""Auditable percent-input analysis for ChIP-qPCR Excel/CSV exports."""

from __future__ import annotations

import argparse
import csv
import math
import re
import statistics
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

ALIASES = {
    "well": {"wellposition", "well", "wellid"},
    "sample": {"samplename", "sample"},
    "target": {"targetname", "target", "assay"},
    "ct": {"ct", "cq", "cqvalue", "ctvalue"},
    "omit": {"omit", "exclude", "excluded"},
    "replicate": {"replicate", "biorep", "biologicalreplicate"},
}
KNOWN_FLAGS = ("OUTLIERRG", "HIGHSD", "THOLDFAIL")


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y"}


def canonical_header(header: Any) -> str | None:
    token = norm(header)
    for key, aliases in ALIASES.items():
        if token in aliases:
            return key
    for flag in KNOWN_FLAGS:
        if token == norm(flag):
            return flag
    return None


def map_headers(headers: list[Any]) -> dict[str, int]:
    mapped = {}
    for index, header in enumerate(headers):
        key = canonical_header(header)
        if key and (key not in mapped or (key == "well" and norm(header) == "wellposition")):
            mapped[key] = index
    return mapped


def valid_mapping(mapped: dict[str, int]) -> bool:
    return {"well", "sample", "ct"}.issubset(mapped)


def load_rows(path: Path, sheet: str | None) -> tuple[str, list[dict[str, Any]]]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            raw = list(csv.reader(handle))
        if not raw:
            raise ValueError("CSV is empty")
        header_index = next((i for i, row in enumerate(raw) if valid_mapping(map_headers(row))), None)
        if header_index is None:
            raise ValueError("Could not find well, sample, and Ct/Cq columns in CSV")
        return path.name, rows_from_table(raw[header_index], raw[header_index + 1 :])

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Input must be .xlsx, .xlsm, or .csv; convert legacy .xls first")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency: install openpyxl") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    candidates = [workbook[sheet]] if sheet else list(workbook.worksheets)
    for worksheet in candidates:
        raw = [list(row) for row in worksheet.iter_rows(values_only=True)]
        for index, row in enumerate(raw):
            if valid_mapping(map_headers(row)):
                return worksheet.title, rows_from_table(row, raw[index + 1 :])
    raise ValueError("Could not find a sheet row containing well, sample, and Ct/Cq columns")


def rows_from_table(headers: list[Any], body: list[list[Any]]) -> list[dict[str, Any]]:
    mapped = map_headers(headers)
    rows = []
    for raw in body:
        item = {key: raw[index] if index < len(raw) else None for key, index in mapped.items()}
        if item.get("well") in (None, "") or item.get("sample") in (None, ""):
            continue
        try:
            item["ct"] = float(item["ct"])
        except (TypeError, ValueError):
            continue
        item["well"] = str(item["well"]).strip()
        item["sample"] = str(item["sample"]).strip()
        item["target"] = str(item.get("target") or "target").strip()
        rows.append(item)
    if not rows:
        raise ValueError("No numeric Ct/Cq data rows found")
    return rows


def replicate_id(row: dict[str, Any], pattern: re.Pattern[str] | None) -> str:
    if row.get("replicate") not in (None, ""):
        return str(row["replicate"]).strip()
    if pattern:
        for value in (row["well"], row["sample"]):
            match = pattern.search(str(value))
            if match:
                return match.group(1) if match.groups() else match.group(0)
        raise ValueError(f"Replicate regex did not match well/sample: {row['well']}")
    match = re.match(r"([A-Za-z]+)", row["well"])
    if not match:
        raise ValueError(f"Cannot infer replicate from well: {row['well']}")
    return match.group(1).upper()


def sample_stats(values: list[float]) -> tuple[float, float | None, float | None]:
    mean = statistics.mean(values)
    if len(values) < 2:
        return mean, None, None
    sd = statistics.stdev(values)
    return mean, sd, sd / math.sqrt(len(values))


def analyze(
    rows: list[dict[str, Any]],
    input_label: str,
    effective_fraction: float,
    exclude_flags: set[str],
    replicate_pattern: re.Pattern[str] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    correction = math.log2(1 / effective_fraction)
    for row in rows:
        row["replicate"] = replicate_id(row, replicate_pattern)
        row["omitted"] = truthy(row.get("omit"))
        row["excluded_by_flag"] = any(truthy(row.get(flag)) for flag in exclude_flags)
        row["included"] = not row["omitted"] and not row["excluded_by_flag"]

    inputs: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    samples: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (row["target"], row["replicate"])
        if row["sample"].casefold() == input_label.casefold():
            inputs[key].append(row)
        else:
            samples[(row["target"], row["sample"], row["replicate"])].append(row)

    calculations = []
    audit = []
    for (target, sample, replicate), ip_rows in sorted(samples.items()):
        input_rows = inputs.get((target, replicate), [])
        usable_input = [r for r in input_rows if r["included"]]
        usable_ip = [r for r in ip_rows if r["included"]]
        if not usable_input or not usable_ip:
            raise ValueError(f"Missing included input or IP wells for {target}/{sample}/{replicate}")
        input_mean = statistics.mean(r["ct"] for r in usable_input)
        ip_mean = statistics.mean(r["ct"] for r in usable_ip)
        adjusted = input_mean - correction
        percent_input = 100 * 2 ** (adjusted - ip_mean)
        calculations.append(
            {
                "target": target,
                "sample": sample,
                "replicate": replicate,
                "input_wells": ",".join(r["well"] for r in usable_input),
                "ip_wells": ",".join(r["well"] for r in usable_ip),
                "n_input": len(usable_input),
                "n_ip": len(usable_ip),
                "input_ct_mean": input_mean,
                "adjusted_input_ct": adjusted,
                "ip_ct_mean": ip_mean,
                "percent_input": percent_input,
                "excluded_wells": ",".join(r["well"] for r in input_rows + ip_rows if not r["included"]),
            }
        )
        for role, group_rows in (("input", input_rows), (sample, ip_rows)):
            for row in group_rows:
                audit.append(
                    {
                        "target": target,
                        "sample_group": sample,
                        "replicate": replicate,
                        "role": role,
                        "well": row["well"],
                        "ct": row["ct"],
                        "included": "Y" if row["included"] else "N",
                        "omit": "Y" if row["omitted"] else "N",
                        **{flag: row.get(flag, "") for flag in KNOWN_FLAGS},
                    }
                )

    grouped: dict[tuple[str, str], list[float]] = defaultdict(list)
    for calc in calculations:
        grouped[(calc["target"], calc["sample"])].append(calc["percent_input"])
    summary = []
    for (target, sample), values in sorted(grouped.items()):
        mean, sd, sem = sample_stats(values)
        summary.append(
            {"target": target, "sample": sample, "n": len(values), "mean_percent_input": mean, "sd": sd, "sem": sem}
        )
    return summary, calculations, audit


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(
    path: Path,
    source: Path,
    sheet: str,
    input_percent: float,
    dilution: float,
    effective_fraction: float,
    exclude_flags: set[str],
    summary: list[dict[str, Any]],
    calculations: list[dict[str, Any]],
    audit: list[dict[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Missing dependency: install openpyxl") from exc
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Summary"
    metadata = [
        ["ChIP-qPCR percent-input summary"],
        ["Source file", str(source.resolve())],
        ["Source sheet", sheet],
        ["Input percent", input_percent],
        ["Additional dilution", dilution],
        ["Effective input fraction", effective_fraction],
        ["Ct correction", math.log2(1 / effective_fraction)],
        ["Excluded flags", ", ".join(sorted(exclude_flags)) or "none"],
        [],
    ]
    for row in metadata:
        ws.append(row)
    append_table(ws, summary)
    for title, data in (("Replicate calculations", calculations), ("Well-level audit", audit)):
        table_ws = workbook.create_sheet(title)
        append_table(table_ws, data)
    prism = workbook.create_sheet("Prism data")
    groups: dict[tuple[str, str], list[tuple[str, float]]] = defaultdict(list)
    for row in calculations:
        groups[(row["target"], row["sample"])].append((row["replicate"], row["percent_input"]))
    for col, ((target, sample), values) in enumerate(sorted(groups.items()), 1):
        prism.cell(1, col, f"{sample} | {target} | %Input")
        for row_index, (_, value) in enumerate(values, 2):
            prism.cell(row_index, col, value)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for column in worksheet.columns:
            letter = column[0].column_letter
            worksheet.column_dimensions[letter].width = min(45, max(10, max(len(str(c.value or "")) for c in column) + 2))
    workbook.save(path)


def append_table(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    ws.append(list(rows[0]))
    for row in rows:
        ws.append(list(row.values()))


def write_plot(path: Path, calculations: list[dict[str, Any]]) -> None:
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError as exc:
        raise RuntimeError("Missing dependency: install matplotlib") from exc
    groups: dict[tuple[str, str], list[float]] = defaultdict(list)
    for row in calculations:
        groups[(row["target"], row["sample"])].append(row["percent_input"])
    labels = [f"{sample}\n{target}" for target, sample in sorted(groups)]
    values = [groups[key] for key in sorted(groups)]
    means = [statistics.mean(group) for group in values]
    errors = [statistics.stdev(group) if len(group) > 1 else 0 for group in values]
    fig, ax = plt.subplots(figsize=(max(5, len(labels) * 1.5), 4.5))
    x = range(len(labels))
    ax.bar(x, means, yerr=errors, width=0.65, color="#74A9CF", edgecolor="#2B5C7B", capsize=4)
    for index, group in enumerate(values):
        offsets = [0] if len(group) == 1 else [(-0.16 + 0.32 * i / (len(group) - 1)) for i in range(len(group))]
        ax.scatter([index + offset for offset in offsets], group, color="#17324D", zorder=3, s=30)
    ax.set_xticks(list(x), labels)
    ax.set_ylabel("% input")
    ax.spines[["top", "right"]].set_visible(False)
    fig.tight_layout()
    fig.savefig(path, dpi=300)
    plt.close(fig)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_file", type=Path)
    parser.add_argument("--sheet", help="Excel sheet name; otherwise auto-detect")
    parser.add_argument("--input-label", default="input")
    parser.add_argument("--input-percent", type=float, required=True, help="Saved input as percent of total chromatin")
    parser.add_argument("--additional-dilution", type=float, default=1.0)
    parser.add_argument("--exclude-flag", action="append", default=[], help="Repeat for OUTLIERRG, HIGHSD, or THOLDFAIL")
    parser.add_argument("--replicate-regex", help="Regex with optional capture group matched against well/sample")
    parser.add_argument("--output-dir", type=Path, default=Path("chip_qpcr_results"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not (0 < args.input_percent <= 100):
        raise ValueError("--input-percent must be >0 and <=100")
    if args.additional_dilution <= 0:
        raise ValueError("--additional-dilution must be >0")
    exclude_flags = {flag.upper() for flag in args.exclude_flag}
    unknown = exclude_flags.difference(KNOWN_FLAGS)
    if unknown:
        raise ValueError(f"Unknown flag(s): {', '.join(sorted(unknown))}")
    effective_fraction = (args.input_percent / 100) / args.additional_dilution
    sheet, rows = load_rows(args.input_file, args.sheet)
    pattern = re.compile(args.replicate_regex) if args.replicate_regex else None
    summary, calculations, audit = analyze(rows, args.input_label, effective_fraction, exclude_flags, pattern)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(args.output_dir / "chip_qpcr_summary.csv", summary)
    write_workbook(
        args.output_dir / "chip_qpcr_analysis.xlsx",
        args.input_file,
        sheet,
        args.input_percent,
        args.additional_dilution,
        effective_fraction,
        exclude_flags,
        summary,
        calculations,
        audit,
    )
    write_plot(args.output_dir / "chip_qpcr_plot.png", calculations)
    print(f"Analyzed {len(rows)} wells into {len(calculations)} biological replicate values")
    print(f"Results: {args.output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (ValueError, RuntimeError, FileNotFoundError, KeyError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(2)
